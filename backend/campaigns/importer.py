"""Excel/CSV import for the FIXED contact schema: name, phone, context, language.

No column mapping — headers are matched case-insensitively. Rows missing a name
or with an unparseable phone are returned in the ``invalid`` bucket so the wizard
can show a valid/invalid preview.
"""
import csv

import phonenumbers
from openpyxl import load_workbook

COLUMNS = ["name", "phone", "context", "language"]


def normalize_phone(raw, region="DE"):
    """Parse a raw phone string to E.164, or return None if invalid."""
    try:
        p = phonenumbers.parse(str(raw).strip(), region)
        if not phonenumbers.is_valid_number(p):
            return None
        return phonenumbers.format_number(p, phonenumbers.PhoneNumberFormat.E164)
    except Exception:
        return None


def _rows(fileobj, filename):
    """Yield dict rows from a .csv or .xlsx file object."""
    if filename.lower().endswith(".csv"):
        lines = (l.decode() if isinstance(l, bytes) else l for l in fileobj)
        return list(csv.DictReader(lines))
    wb = load_workbook(fileobj, read_only=True, data_only=True)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        # skip empty header columns so blank headers can't clobber real keys
        rows.append({h: c.value for h, c in zip(header, row) if h})
    return rows


def parse_contacts(fileobj, filename, region="DE"):
    """Returns (valid: list[dict], invalid: list[dict])."""
    valid, invalid = [], []
    for r in _rows(fileobj, filename):
        # normalise header casing/whitespace for robustness
        r = {(k or "").strip().lower(): v for k, v in r.items()}
        name = (r.get("name") or "").strip()
        phone = normalize_phone(r.get("phone"), region)
        if not name or not phone:
            invalid.append({**r, "_error": "missing name or invalid phone"})
            continue
        valid.append(
            {
                "name": name,
                "phone": phone,
                "context": (r.get("context") or "").strip(),
                "language": (r.get("language") or "en").strip() or "en",
            }
        )
    return valid, invalid
